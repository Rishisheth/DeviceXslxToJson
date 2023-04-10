import json
import sys
from openpyxl import load_workbook

def xlsx_to_json(file_path):
    wb = load_workbook(file_path)
    all_sheets = wb.sheetnames

    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        data = []

        # Get headers from the first row
        headers = [cell.value for cell in ws[1]]

        # Iterate through the rows and create JSON objects
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for key, value in zip(headers, row):
                row_data[key] = value
            data.append(row_data)

        output_file = "{}.json".format(sheet_name)
        with open(output_file, "w") as f:
            json.dump(data, f, indent=4)

        print("Generated JSON file: {}".format(output_file))

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Please provide a file path to the xlsx file.")
        sys.exit(1)

    file_path = sys.argv[1]
    xlsx_to_json(file_path)